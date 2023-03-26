
from openpyxl import Workbook
from openpyxl.styles import Font
from PyPDF2 import PdfReader
from datetime import datetime
import re
import fileinput
import io
from distutils.log import debug
from fileinput import filename
from flask import *
app = Flask(__name__)


# Excel sheet creation
workbook = Workbook()
sheet = workbook.active

dates_list = []  # the list of all the dates of assignments
info_list = []  # the list of all the assignments

# this was done with two months list because it makes it easier to replace text with dates
months_list = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October",
               "November", "December"]
shortened_months_list = ["Jan", "Feb", "Mar", "Apr",
    "May", "June", "July", "Aug", "Sep", "Oct", "Nov", "Dec"]

# temporary list of inputs to program
file_list = []
# lets us adjust color
color_list = ["00008000", "00003366", "00FF9900", "00CCCCFF", "00FF8080"]

total_input = 0   # tracks how many items have been put in the spreadsheet
class_count = 0  # keeps track of how many classes we have inputted data for


def get_data(filename):
    """
    This function takes in the name of a file, either pdf or plaintext for now, and then reads it to find all due
    dates. If the file it's given is a pdf, it converts it to plaintext using PyPDF2. Otherwise, it just reads the file
    and uses regular expressions to find the information we need.
    :param filename: the name of the file that is going to be read
    """
    global dates_list
    global info_list
    global class_count
    count_entries = 0  # used to prevent duplicate entries to lists

    if filename[len(filename) - 4:len(filename)] == ".pdf":
        reader = PdfReader(filename)
        number_of_pages = len(reader.pages)
        # write to tempt.txt, will overwrite any file with the same name in project
        file1 = open("temp2.txt", "w")

        for j in range(0, number_of_pages):  # copy each page of PDF
            page = reader.pages[j]
            text = page.extract_text()
            file1.writelines(text)
            file1.writelines("\n")

        file1.close()  # close file
        filename = "temp2.txt"
        # reads our file and adds a line break for every period, used to fix issue where conversion from pdf to
        with fileinput.FileInput(filename, inplace=True, backup='.bak') as new_file:
            for line in new_file:
                print(line.replace(". ", "\n"), end='')

    textfile = open(filename, 'r')
    filetext = textfile.read()
    textfile.close()
    final_matches = re.findall(r"([A-Z][a-z]+)(.)([0-9]+)(.+)", filetext)

    for match in final_matches:  # this is where you work through all your data to get rid of bad data
        # temp_list[0] = Month
        # temp_list[1] = " "
        # temp_list[2] = Day
        # temp_list[3] = Information
        temp_list = list(match)

        if match[0] in months_list or match[0] in shortened_months_list:
            if len(temp_list[3]) > 1:
                if temp_list[3][0] == " ":
                    temp_list[3] = temp_list[3][1:len(temp_list[3])]
                if not temp_list[3][0].islower():
                    for y in range(0, len(months_list)):
                        temp_list[0] = temp_list[0].replace(
                            months_list[y], str(y + 1) + "/", )
                        if count_entries == 0 and "/" in temp_list[0]:
                            dates_list.append(temp_list[0] + temp_list[2])
                            info_list.append(temp_list[3])
                            count_entries += 1
                    for z in range(0, len(shortened_months_list)):
                        temp_list[0] = temp_list[0].replace(
                            shortened_months_list[z], str(z + 1) + "/", )
                        if count_entries == 0 and "/" in temp_list[0]:
                            dates_list.append(temp_list[0] + temp_list[2])
                            info_list.append(temp_list[3])
                            count_entries += 1
        count_entries = 0
    x = 0
    while x < len(info_list):
        info_list[x] = info_list[x].replace("| ", "")
        if info_list[x][0] == ":":
            info_list[x] = info_list[x][1:len(info_list[x])]
        if "No Class" in info_list[x] or "No class" in info_list[x] or "no class" in info_list[x] \
            or "Reading Day" in info_list[x] or "Reading day" in info_list[x] or "reading day" in info_list[x]:
            del info_list[x]
            del dates_list[x]
        x += 1
    fill_spreadsheet(dates_list, info_list)
    class_count += 1
    dates_list = []
    info_list = []

    return


def fill_spreadsheet(list_dates, list_info):
    """
    This function takes in the two lists that are created in get_data and populates an excel spreadsheet with the
    information, in the order that the syllabi were read.
    :param list_dates: the list of all the dates found on the syllabi
    :param list_info:  the list of all the assignments found on the syllabi
    """
    global total_input
    global class_count
    temp_addition = 0

    # creates headers for each column
    sheet["A1"] = "Date"
    sheet["B1"] = "Assignment"

    for j in range(0, len(list_dates)):
        # fills first column with first names
        placeholder = "A" + str(j + 2 + total_input)
        sheet[placeholder] = list_dates[j]
        temp_addition += 1

    for m in range(0, len(list_dates)):
        # fills second column with last names
        placeholder = "B" + str(m + 2 + total_input)
        sheet[placeholder] = list_info[m]

    for d in range(0, len(list_dates)):
        # fills third column with class number (helps us with coloring later on)
        placeholder = "C" + str(d + 2 + total_input)
        sheet[placeholder] = class_count

    total_input += temp_addition
    workbook.save(filename="your_new_schedule.xlsx")
    return


def sort_spreadsheet():
    """
    This function is called after the spreadsheet has been created. It sorts puts the cell entries in chronological
    order and colors them according to class.
    """
    sheet_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        sheet_data.append(row)

    sheet_data.sort(key=lambda row: datetime.strptime(row[0], '%m/%d'))
    sheet.delete_rows(2, sheet.max_row-1)

    for row in sheet_data:
        sheet.append(row)

    # change the color of the important cells and delete the last column
    for cell in range(0, sheet.max_row-1):
        col1 = "A" + str(cell + 2)
        col2 = "B" + str(cell + 2)
        sheet[col1].font = Font(color=color_list[(sheet_data[cell][2] % 5)])
        sheet[col2].font = Font(color=color_list[(sheet_data[cell][2] % 5)])
        sheet.delete_cols(3)

    workbook.save(filename="your_new_schedule.xlsx")
    return


@app.route('/')
def main():
    return render_template("index.html")


@app.route('/upload', methods=['POST'])
def success():
    if request.method == 'POST':
        files = request.files.getlist('file')
        for f in files:
            f.save(f.filename)  # saves the file
            file_list.append(f.filename)  # adds to the list
        for file in file_list:  # this modifies the file. The sort_spreadsheet modifies the workbook. We want to return the workbook
             get_data(file)
        sort_spreadsheet()
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Serve the buffer as a file download
    return send_file(
        buffer,
        as_attachment=True,
        attachment_filename='output.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
   
if __name__ == '__main__':  
    app.run(debug=True)

